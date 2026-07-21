"""
============================================================
سكريبت التحديث الكامل لداشبورد التحصيلات - نسخة الإنتاج
============================================================
يعمل بالكامل من غير أي تدخل بشري:
  1) يحمّل 4 ملفات من مجلد Google Drive (مطابقة مرنة للأسماء)
  2) يبني RAWDATA و PERIOD_DATA (منطق v8 المُختبَر)
  3) يشغّل قائمة الفحص الآلية
  4) يستبدل البيانات في dashboard_collections.html عبر raw_decode
  5) يتحقق من صحة JavaScript
  6) عند النجاح: يكتب الملف الجديد في مكانه بالريبو (ليتم الـ commit بعده)
  7) عند أي فشل: يرسل إيميل تنبيه بالتفاصيل ولا يكتب أي شيء
============================================================
"""
import openpyxl, pandas as pd, json, re, sys, subprocess, os, tempfile, traceback

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from drive_downloader import download_all_required_files, norm_ar
from email_alert import send_alert_email

# ============================================================
# 0) الإعدادات من متغيرات البيئة (تُمرَّر من GitHub Actions Secrets)
# ============================================================
GDRIVE_KEY_PATH   = os.environ['GDRIVE_KEY_PATH']
GDRIVE_FOLDER_ID  = os.environ['GDRIVE_FOLDER_ID']
DASHBOARD_PATH    = os.environ.get('DASHBOARD_PATH', 'dashboard_collections.html')
EMAIL_FROM        = os.environ['ALERT_EMAIL_FROM']
EMAIL_APP_PASS    = os.environ['ALERT_EMAIL_APP_PASSWORD']
EMAIL_TO          = os.environ['ALERT_EMAIL_TO']
YEAR              = int(os.environ.get('DASHBOARD_YEAR', '2026'))
DUE_TOLERANCE     = 5

ALERT_LOG, WARNING_LOG = [], []
def fatal(msg): ALERT_LOG.append(msg)
def warn(msg): WARNING_LOG.append(msg)

def die_with_alert(reason_lines):
    send_alert_email(
        subject="🛑 توقف تحديث داشبورد التحصيلات - يحتاج مراجعة",
        body_lines=["السكريبت توقف ولم يُنشَر أي تحديث. الأسباب:", ""] + reason_lines,
        to_addr=EMAIL_TO, from_addr=EMAIL_FROM, app_password=EMAIL_APP_PASS
    )
    print("\n🛑 توقف — تم إرسال تنبيه بريدي. الأسباب:")
    for r in reason_lines: print("  -", r)
    sys.exit(1)

# ============================================================
# 1) تحميل الملفات من Google Drive
# ============================================================
print("=== تحميل الملفات من Google Drive ===")
tmp_dir = tempfile.mkdtemp(prefix='dashboard_input_')
try:
    file_paths, mtimes = download_all_required_files(GDRIVE_KEY_PATH, GDRIVE_FOLDER_ID, tmp_dir)
    print(f"  ✅ تم تحميل 4 ملفات: {list(file_paths.keys())}")
    for k, v in mtimes.items():
        print(f"     {k}: آخر تعديل {v}")
except Exception as e:
    die_with_alert([f"فشل تحميل الملفات من Google Drive: {e}"])

PATHS = {
    'main_file': file_paths['main_file'],
    'revenues':  file_paths['revenues'],
    'advance':   file_paths['advance'],
    'deduct':    file_paths['deduct'],
}

# ============================================================
# 2) الثوابت (من قسم 15 من v8)
# ============================================================
REAL_METHODS = {'أونلاين','تحويل بنكى','شيك','كي نت','نقدا','مستحق من أطراف ذي صلة'}
REAL_TYPES   = {'دفعة ايجار','إيرادات أخري'}
PORT_SHORTS  = {
    '1 - أسواق القرين':     'القرين',
    '2 - مجمعات العقيلة':   'العقيلة',
    '3 - أسواق العارضية':   'العارضية',
    '4 - العقارات السكنية':  'السكني',
    '5 - العقارات التجارية': 'التجاري',
    '8 - أملاك خاصة':       'أملاك خاصة',
}
PORTS = list(PORT_SHORTS.keys())
PCT_METHODS = ['أونلاين','تحويل بنكى','شيك','كي نت','نقدا']
METHODS_WITH_BREAKDOWN = [
    'أونلاين','تحويل بنكى','شيك','كي نت','نقدا','مستحق من أطراف ذي صلة',
    'اعفاء','إعفاء قانوني','إعفاء إداري','شيكات مؤجلة'
]
RAWDATA_EXEMPTION_FILTER = {'اعفاء', 'إعفاء قانوني'}
APPROVED_STATUS = 'معتمد'
DEDUCT_PORT_MAP = {
    'العارضية': '3 - أسواق العارضية', 'العقيلة': '2 - مجمعات العقيلة',
    'القرين': '1 - أسواق القرين', 'تجاري': '5 - العقارات التجارية',
    'سكني': '4 - العقارات السكنية', 'التجاري': '5 - العقارات التجارية',
    'السكني': '4 - العقارات السكنية', 'أملاك خاصة': '8 - أملاك خاصة',
    'املاك خاصة': '8 - أملاك خاصة',
}
KNOWN_METHODS = set(METHODS_WITH_BREAKDOWN) | {
    'تحويل من دفعة تحت الحساب','تحويل تأمين لإيجار',
    'ايداعات بالمحكمة لم تحصل','تسويات','إقرار دين','حسابات معلقة'
}
KNOWN_STATUSES = {'معتمد', 'غير معتمد'}

try:
    # ============================================================
    # 3) قراءة الملف الأساسي + بناء RAWDATA
    # ============================================================
    print("=== قراءة الملف الأساسي ===")
    wb = openpyxl.load_workbook(PATHS['main_file'], read_only=True)
    ws = wb['sheet1']

    all_rows = []
    filtered_count, filtered_value = 0, 0
    unknown_methods, unknown_statuses = set(), set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        portfolio_group = row[21]
        if portfolio_group and 'دبى' in str(portfolio_group):
            continue
        col_date = row[7]
        if not col_date or not hasattr(col_date, 'month'):
            continue
        method = str(row[8]).strip() if row[8] else ''
        status = str(row[9]).strip() if row[9] else ''

        if method and method not in KNOWN_METHODS:
            unknown_methods.add(method)
        if status and status not in KNOWN_STATUSES:
            unknown_statuses.add(status)

        if method in RAWDATA_EXEMPTION_FILTER and status != APPROVED_STATUS:
            filtered_count += 1
            filtered_value += float(row[16] or 0)
            continue

        p_long = str(portfolio_group) if portfolio_group else ''
        all_rows.append({
            'p': p_long, 'p_short': PORT_SHORTS.get(p_long, ''),
            'c': str(row[2]) if row[2] else '', 'tn': str(row[10]).strip() if row[10] else '',
            'cm': col_date.month, 'cy': col_date.year,
            'dm': row[15].month if (row[15] and hasattr(row[15], 'month')) else None,
            'dy': row[15].year if (row[15] and hasattr(row[15], 'year')) else None,
            'm': method, 't': str(row[5]).strip() if row[5] else '',
            'v': float(row[16] or 0), 'x': str(row[23]).strip() if row[23] else ''
        })
    wb.close()

    if unknown_methods:
        fatal(f"طرق دفع غير معروفة ظهرت في الملف: {unknown_methods}")
    if unknown_statuses:
        fatal(f"قيم حالة غير معروفة ظهرت في الملف: {unknown_statuses}")

    MONTHS = sorted(set(r['cm'] for r in all_rows if r['cm'] and r['cy'] == YEAR))
    if not MONTHS:
        fatal("لم يُعثر على أي بيانات للسنة المحددة في الملف الأساسي")

    print(f"  صفوف صالحة: {len(all_rows):,} | إعفاءات غير معتمدة مُستبعدة: {filtered_count}")
    print(f"  الأشهر المكتشفة: {MONTHS}")

    if ALERT_LOG:
        die_with_alert(ALERT_LOG)

    def due_key(r):
        if not r['dy']: return None
        if r['dy'] < YEAR: return 'b'
        if r['dy'] > YEAR: return 'a'
        if r['dm'] is None: return None
        return f"{r['dm']:02d}"

    def bucket_sum(rlist):
        b = {}
        for r in rlist:
            k = due_key(r)
            if k is None: continue
            b[k] = b.get(k, 0) + r['v']
        return {k: round(v) for k, v in b.items() if round(v) != 0}

    def aggregate_by_contract(rows_subset):
        d = {}
        for r in rows_subset:
            c = r['c']
            if not c: continue
            if c not in d:
                d[c] = {'t': r['tn'], 'v': 0, 'p': r['p_short'], 'months': set()}
            d[c]['v'] += r['v']
            if r['tn'] and not d[c]['t']: d[c]['t'] = r['tn']
            if r['dy'] and r['dm']: d[c]['months'].add(f"{r['dy']:04d}-{r['dm']:02d}")
        return [{'c': c, 't': i['t'], 'v': round(i['v']), 'p': i['p'], 'months': sorted(i['months'])}
                for c, i in sorted(d.items(), key=lambda x: -x[1]['v']) if round(i['v']) > 0]

    def compute_rawdata(rows, col_month):
        rows = [r for r in rows if (r['cm']==col_month if col_month is not None else True) and r['cy']==YEAR]
        ins_rows = [r for r in rows if r['t']=='دفعة تأمين']
        sub_acc_rows = [r for r in rows if r['t']=='دفعة تحت الحساب']
        special_rows = [r for r in rows if r['t']=='دفعة تاجير مخصص']
        real_rows = [r for r in rows if r['m'] in REAL_METHODS and r['t'] in REAL_TYPES]
        svc_rows = [r for r in real_rows if r['t']=='إيرادات أخري']
        ex_rows = [r for r in rows if r['m'] in {'اعفاء','إعفاء قانوني'}]
        lex_rows = [r for r in rows if r['m']=='إعفاء قانوني']
        chk_rows = [r for r in rows if r['m']=='شيكات مؤجلة']
        total_cash_rows = [r for r in rows if r['m'] in REAL_METHODS]
        ex_regular_rows = [r for r in rows if r['m']=='اعفاء']
        ex_legal_rows = [r for r in rows if r['m']=='إعفاء قانوني']
        suspense_rows = [r for r in rows if r['m']=='تحويل من دفعة تحت الحساب']
        ins_tr_rows = [r for r in rows if r['m']=='تحويل تأمين لإيجار']
        court_rows = [r for r in rows if r['m']=='ايداعات بالمحكمة لم تحصل']
        settle_rows = [r for r in rows if r['m']=='تسويات']
        debt_ack_rows = [r for r in rows if r['m']=='إقرار دين']
        pending_rows = [r for r in rows if r['m']=='حسابات معلقة']

        method_data = {m: {'total': round(sum(r['v'] for r in rows if r['m']==m)),
                            'due': bucket_sum([r for r in rows if r['m']==m])}
                       for m in METHODS_WITH_BREAKDOWN}
        chk_by_type = {
            'rentals_services': round(sum(r['v'] for r in chk_rows if r['t'] in REAL_TYPES)),
            'insurance': round(sum(r['v'] for r in chk_rows if r['t']=='دفعة تأمين')),
            'sub_account': round(sum(r['v'] for r in chk_rows if r['t']=='دفعة تحت الحساب')),
        }
        svc_by_type = {}
        for r in svc_rows: svc_by_type[r['x']] = svc_by_type.get(r['x'],0)+r['v']
        svc_by_type = {k: round(v) for k,v in sorted(svc_by_type.items(), key=lambda x:-x[1]) if round(v)>0}
        contract_counts = {m: len(set(r['c'] for r in rows if r['m']==m and r['c'])) for m in PCT_METHODS}

        return {
            'real': {'total': round(sum(r['v'] for r in real_rows)), 'due': bucket_sum(real_rows)},
            'svc': {'total': round(sum(r['v'] for r in svc_rows)), 'due': bucket_sum(svc_rows)},
            'ex': {'total': round(sum(r['v'] for r in ex_rows)), 'due': bucket_sum(ex_rows)},
            'lex': {'total': round(sum(r['v'] for r in lex_rows)), 'due': bucket_sum(lex_rows)},
            'chk': {'total': round(sum(r['v'] for r in chk_rows)), 'due': bucket_sum(chk_rows)},
            'ins': round(sum(r['v'] for r in ins_rows)), 'ins_tr': round(sum(r['v'] for r in ins_tr_rows)),
            'court': round(sum(r['v'] for r in court_rows)), 'settle': round(sum(r['v'] for r in settle_rows)),
            'suspense': round(sum(r['v'] for r in suspense_rows)), 'debt_ack': round(sum(r['v'] for r in debt_ack_rows)),
            'pending': round(sum(r['v'] for r in pending_rows)),
            'suspense_due': bucket_sum(suspense_rows), 'ins_tr_due': bucket_sum(ins_tr_rows),
            'court_due': bucket_sum(court_rows), 'settle_due': bucket_sum(settle_rows),
            'debt_ack_due': bucket_sum(debt_ack_rows), 'pending_due': bucket_sum(pending_rows),
            'sub_acc': round(sum(r['v'] for r in sub_acc_rows)), 'special': round(sum(r['v'] for r in special_rows)),
            'total_cash': round(sum(r['v'] for r in total_cash_rows)),
            'methods': method_data, 'chk_by_type': chk_by_type, 'svc_types': svc_by_type,
            'contract_counts': contract_counts,
            'top_ex': aggregate_by_contract(ex_regular_rows), 'top_lex': aggregate_by_contract(ex_legal_rows),
            'grand': round(sum(r['v'] for r in rows))
        }

    print("=== بناء RAWDATA ===")
    rawdata_output = {}
    for port in PORTS:
        port_rows = [r for r in all_rows if r['p']==port]
        rawdata_output[port] = {str(m): compute_rawdata(port_rows, m) for m in MONTHS}
        rawdata_output[port]['all'] = compute_rawdata(port_rows, None)
    rawdata_output['all'] = {str(m): compute_rawdata(all_rows, m) for m in MONTHS}
    rawdata_output['all']['all'] = compute_rawdata(all_rows, None)
    print(f"  grand: {rawdata_output['all']['all']['grand']:,} د.ك")

    # ============================================================
    # 4) بناء PERIOD_DATA
    # ============================================================
    print("=== بناء PERIOD_DATA ===")
    wb_rev = openpyxl.load_workbook(PATHS['revenues'], read_only=True)
    rev_sheet_name = wb_rev.sheetnames[0]
    ws_rev = wb_rev[rev_sheet_name]
    rev_rows_raw = list(ws_rev.iter_rows(min_row=1, max_row=14, values_only=True))
    header_row_idx = next((i for i, r in enumerate(rev_rows_raw) if r[2] and 'يناير' in str(r[2])), None)
    if header_row_idx is None:
        fatal("لم يُعثر على صف الهيدر (يناير) في شيت الإيرادات الشهرية")

    short_to_long = {norm_ar(s): l for l, s in PORT_SHORTS.items()}
    revenues = {}
    if header_row_idx is not None:
        for r in rev_rows_raw[header_row_idx+1:]:
            label = r[1]
            if not label: continue
            label_norm = norm_ar(label)
            if label_norm not in short_to_long: continue
            long_p = short_to_long[label_norm]
            revenues[long_p] = {mi: (float(r[ci]) if isinstance(r[ci], (int,float)) else 0.0)
                                 for mi, ci in enumerate(range(2, 2+len(MONTHS)), start=1)}
    wb_rev.close()
    missing_ports = [p for p in PORTS if p not in revenues]
    if missing_ports:
        fatal(f"محافظ غير مُطابَقة في شيت الإيرادات الشهرية: {missing_ports}")

    if ALERT_LOG:
        die_with_alert(ALERT_LOG)

    wb2 = openpyxl.load_workbook(PATHS['main_file'], read_only=True)
    ws2 = wb2['sheet1']
    main_paid = {}
    for row in ws2.iter_rows(min_row=2, values_only=True):
        pg = row[21]
        if not pg or 'دبى' in str(pg): continue
        if str(row[5]).strip() not in REAL_TYPES: continue
        if str(row[9]).strip() != APPROVED_STATUS: continue
        dd = row[15]
        if not dd or not hasattr(dd, 'month'): continue
        if not (dd.year==YEAR and 1<=dd.month<=len(MONTHS)): continue
        key = (str(pg), dd.month)
        main_paid[key] = main_paid.get(key, 0) + float(row[16] or 0)
    wb2.close()

    df_adv = pd.read_excel(PATHS['advance'])
    df_adv = df_adv[df_adv['نوع الدفعة'].isin(REAL_TYPES)]
    df_adv = df_adv[~df_adv['مجموعة المحفظة'].astype(str).str.contains('دبى', na=False)]
    df_adv['due_date'] = pd.to_datetime(df_adv['تاريخ الاستحقاق'], errors='coerce')
    df_adv = df_adv.dropna(subset=['due_date'])
    df_adv = df_adv[(df_adv['due_date'].dt.year==YEAR) & (df_adv['due_date'].dt.month.between(1,len(MONTHS)))]
    adv_paid = {}
    for _, row in df_adv.iterrows():
        key = (str(row['مجموعة المحفظة']), row['due_date'].month)
        adv_paid[key] = adv_paid.get(key, 0) + float(row['قيمة الدفعة'] or 0)

    df_deduct = pd.read_excel(PATHS['deduct'])
    df_deduct['due_date'] = pd.to_datetime(df_deduct['تاريخ الاستحقاق'], errors='coerce')
    df_deduct = df_deduct.dropna(subset=['due_date'])
    deduct_paid = {}
    for _, row in df_deduct.iterrows():
        if row['due_date'].year != YEAR: continue
        month = row['due_date'].month
        if not (1<=month<=len(MONTHS)): continue
        pr = str(row['مجموعة المحفظة']).strip() if row['مجموعة المحفظة'] else ''
        pl = DEDUCT_PORT_MAP.get(pr) or DEDUCT_PORT_MAP.get(norm_ar(pr))
        if not pl: continue
        key = (pl, month)
        deduct_paid[key] = deduct_paid.get(key, 0) + float(row['مدين'] or 0)

    def compute_period(port_list, months_list):
        rev = sum(revenues[p][m] for p in port_list for m in months_list if p in revenues and m in revenues[p])
        paid = (sum(main_paid.get((p,m),0) for p in port_list for m in months_list) +
                sum(adv_paid.get((p,m),0) for p in port_list for m in months_list) +
                sum(deduct_paid.get((p,m),0) for p in port_list for m in months_list))
        unpaid = rev - paid
        rate = round((unpaid/rev*100), 2) if rev else 0
        return {'rev': round(rev,3), 'paid': round(paid,3), 'unpaid': round(unpaid,3), 'rate': rate}

    period_data_output = {'all': {str(m): compute_period(PORTS, [m]) for m in MONTHS}}
    period_data_output['all']['all'] = compute_period(PORTS, MONTHS)
    for port in PORTS:
        period_data_output[port] = {str(m): compute_period([port], [m]) for m in MONTHS}
        period_data_output[port]['all'] = compute_period([port], MONTHS)
    print(f"  rev: {period_data_output['all']['all']['rev']:,.0f} | paid: {period_data_output['all']['all']['paid']:,.0f}")

    # ============================================================
    # 5) قائمة الفحص الآلية
    # ============================================================
    print("=== قائمة الفحص الآلية ===")
    d_all = rawdata_output['all']['all']
    methods_sum = sum(v['total'] for v in d_all['methods'].values())
    totals6_sum = d_all['suspense'] + d_all['ins_tr'] + d_all['court'] + d_all['settle'] + d_all['debt_ack'] + d_all['pending']
    grand_diff = abs((methods_sum + totals6_sum) - d_all['grand'])
    if grand_diff > DUE_TOLERANCE:
        fatal(f"فحص الإجمالي فشل: فرق {grand_diff:,} > {DUE_TOLERANCE}")

    chk_check = d_all['chk_by_type']['rentals_services'] + d_all['chk_by_type']['insurance'] + d_all['chk_by_type']['sub_account']
    if abs(chk_check - d_all['chk']['total']) > DUE_TOLERANCE:
        fatal(f"فحص chk_by_type فشل")

    for port in PORTS:
        if port not in rawdata_output:
            fatal(f"محفظة مفقودة من RAWDATA: {port}")

    rev_all_check = sum(period_data_output[p]['all']['rev'] for p in PORTS)
    if abs(rev_all_check - period_data_output['all']['all']['rev']) > 2:
        fatal(f"فحص rev الكلي فشل")

    if ALERT_LOG:
        die_with_alert(ALERT_LOG)
    print("  ✅ كل الفحوصات نجحت")

    # ============================================================
    # 6) الاستبدال الجراحي في HTML
    # ============================================================
    print("=== الاستبدال في HTML ===")
    with open(DASHBOARD_PATH, 'r', encoding='utf-8') as f:
        html_content = f.read()

    decoder = json.JSONDecoder()
    def replace_json_var(content, var_name, new_obj):
        marker = f'const {var_name} = '
        start = content.find(marker)
        if start == -1:
            marker = f'const {var_name} ='
            start = content.find(marker)
            if start == -1:
                fatal(f"لم يُعثر على متغير {var_name} في ملف HTML")
                return content
        json_start = start + len(marker)
        while content[json_start] in ' \n\t':
            json_start += 1
        obj, end_idx = decoder.raw_decode(content, json_start)
        semi_idx = content.find(';', end_idx)
        new_json_str = json.dumps(new_obj, ensure_ascii=False, separators=(',', ':'))
        return content[:json_start] + new_json_str + content[semi_idx:]

    new_html = replace_json_var(html_content, 'RAWDATA', rawdata_output)
    new_html = replace_json_var(new_html, 'PERIOD_DATA', period_data_output)

    if ALERT_LOG:
        die_with_alert(ALERT_LOG)

    if 'const PORTS' not in new_html or 'const SHORTS' not in new_html:
        fatal("const PORTS أو const SHORTS مفقودان من الملف الناتج")

    scope_match = re.search(r"getScope\(\)\{[^}]*return\[([0-9,]+)\]", new_html)
    if scope_match:
        scope_nums = [int(x) for x in scope_match.group(1).split(',')]
        if scope_nums != MONTHS:
            warn(f"getScope() تُرجع {scope_nums} لكن الأشهر الفعلية {MONTHS} — يحتاج تحديث يدوي (خطأ 33)")

    if ALERT_LOG:
        die_with_alert(ALERT_LOG)

    print("=== فحص صياغة JavaScript ===")
    scripts = re.findall(r'<script>(.*?)</script>', new_html, re.DOTALL)
    for i, s in enumerate(scripts):
        tmp_js = os.path.join(tmp_dir, f'_check_{i}.js')
        with open(tmp_js, 'w', encoding='utf-8') as f:
            f.write(s)
        result = subprocess.run(['node', '--check', tmp_js], capture_output=True, text=True)
        if result.returncode != 0:
            fatal(f"خطأ صياغة JavaScript: {result.stderr}")

    if ALERT_LOG:
        die_with_alert(ALERT_LOG)

    with open(DASHBOARD_PATH, 'w', encoding='utf-8') as f:
        f.write(new_html)

    print("\n✅✅✅ نجح التحديث بالكامل")

    if WARNING_LOG:
        send_alert_email(
            subject="⚠️ تحديث داشبورد التحصيلات نجح لكن به ملاحظات",
            body_lines=["التحديث تم بنجاح ونُشر، لكن توجد ملاحظات تحتاج مراجعتك:", ""] + WARNING_LOG,
            to_addr=EMAIL_TO, from_addr=EMAIL_FROM, app_password=EMAIL_APP_PASS
        )

except SystemExit:
    raise
except Exception as e:
    die_with_alert([f"خطأ غير متوقع: {e}", "", traceback.format_exc()])
